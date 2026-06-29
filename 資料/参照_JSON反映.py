#!/usr/bin/env python3
"""
xlsx の「🔗 参照」列、および 概要/名前/ふりがな/役 などのテキスト欄を
japan_history_data.json の各エンティティに反映する。

使い方:
    # ドライラン（差分プレビューのみ、ファイル変更なし）
    python3 資料/参照_JSON反映.py

    # 実際に反映（自動バックアップ + version bump）
    python3 資料/参照_JSON反映.py --write

    # version bump を抑制したい場合
    python3 資料/参照_JSON反映.py --write --no-version-bump

反映対象カラム:
    人物系シート: 名前(n) / ふりがな(fg) / 役/分類(role) / 概要(desc) / 🔗 参照(refs)
    カテゴリ:     名称(n) / ふりがな(fg) / 概要(desc) / 🔗 参照(refs)
    能曲:        曲名(n) / ふりがな(fg) / 分類(cls) / シテ類型 / シテ / ワキ / 出典 / 時代 /
                 作者 / 宗教 / 注記(notes) / 🔗 参照(refs)

JSON への書き込み形式（refs）:
    "refs": [
      {"id": "saicho", "role": "開祖"},
      {"id": "ennin"}                  # role が空のときは role キーを省略
    ]
"""
import json, re, os, sys, argparse, shutil, datetime
from openpyxl import load_workbook

ROOT = '/Users/TN/Downloads/日本歴史年表'
JSON_PATH = os.path.join(ROOT, 'japan_history_data.json')
XLSX_PATH = os.path.join(ROOT, '資料/全データ一覧_点検用.xlsx')

TOKEN_RE = re.compile(r'^\s*([^()（）;；]+?)\s*(?:[（(]\s*([^)）]*?)\s*[)）])?\s*$')
ID_PATTERN = re.compile(r'^[A-Za-z0-9_]+$')

# ----- シート種別ごとに「ヘッダ表記 → JSON フィールド名」のマップ -----
# ※ 年(s/e) や 時代(era for person/cat) は表示用に整形されるため反映対象外。
FIELD_MAPS = {
    'person': {'名前':'n', 'ふりがな':'fg', '役/分類':'role', '概要':'desc'},
    'cat':    {'名称':'n', 'ふりがな':'fg', '概要':'desc'},
    'noh':    {'曲名':'n', 'ふりがな':'fg', '分類':'cls', 'シテ類型':'shite_type',
               'シテ':'shite', 'ワキ':'waki', '出典':'source', '時代':'era',
               '作者':'author', '宗教':'religion', '注記':'notes'},
}

def _detect_sheet_type(headers):
    """シートのヘッダ行からシート種別を判定。person / cat / noh / None。"""
    if 'ピン数' in headers: return 'person'
    if '曲名'   in headers: return 'noh'
    if '名称'   in headers and '開始' in headers: return 'cat'
    return None

def parse_refs(text, gid, name_to_ids):
    """`token1(役割1); token2; token3(役割3)` → [{'id':..., 'role':...?}]
    token は ID（saicho）または 名前（最澄）どちらでも可。
    エンティティ名は全カテゴリ横断で `n` フィールドと一致させる。
    """
    out = []
    for tok in re.split(r'[;；]', text or ''):
        s = tok.strip()
        if not s: continue
        m = TOKEN_RE.match(s)
        if not m:
            raise ValueError(f'書式不正なトークン: `{s}`')
        raw, role = m.group(1).strip(), (m.group(2) or '').strip()
        # 1) ID として存在すれば採用
        if raw in gid:
            resolved = raw
        # 2) 名前として一意なら採用
        elif raw in name_to_ids:
            ids = name_to_ids[raw]
            if len(ids) == 1:
                resolved = ids[0][0]
            else:
                choices = ', '.join(f'{i}({lb})' for i, lb in ids)
                raise ValueError(f'名前重複: `{raw}` → {choices} のいずれか')
        # 3) ASCII っぽければ ID 不在
        elif ID_PATTERN.match(raw):
            raise ValueError(f'ID不在: `{raw}`')
        # 4) 名前未一致
        else:
            sugg = [nm for nm in name_to_ids if raw in nm or nm in raw][:3]
            hint = f'（近似: {", ".join(sugg)}）' if sugg else ''
            raise ValueError(f'名前未一致: `{raw}`{hint}')
        ref = {'id': resolved}
        if role: ref['role'] = role
        out.append(ref)
    return out

def main():
    ap = argparse.ArgumentParser(description='xlsx の 参照 を JSON に反映する')
    ap.add_argument('--write', action='store_true', help='実際にファイルを書き換える（既定はドライラン）')
    ap.add_argument('--no-version-bump', action='store_true', help='version を上げない')
    args = ap.parse_args()

    if not os.path.exists(XLSX_PATH):
        print(f'ERROR: xlsx が見つかりません: {XLSX_PATH}', file=sys.stderr); sys.exit(2)
    if not os.path.exists(JSON_PATH):
        print(f'ERROR: JSON が見つかりません: {JSON_PATH}', file=sys.stderr); sys.exit(2)

    # JSON ロード
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        d = json.load(f)

    # グローバル ID／名前 索引（参照先存在チェック・名前解決用）
    gid = {}
    name_to_ids = {}
    def reg(arr, label):
        for p in arr:
            if p.get('id'):
                gid[p['id']] = (p.get('n', ''), label)
                nm = p.get('n', '')
                if nm:
                    name_to_ids.setdefault(nm, []).append((p['id'], label))
    people_cats = [('persons','人物'),('emperors','天皇'),('rulers','為政者'),
                   ('religious','宗教者'),('pms','総理大臣'),('rakugo','落語家'),
                   ('kodan','講談師')]
    for k, lb in people_cats: reg(d[k], lb)
    reg(d['noh'], '能曲')
    for c in d['cats']: reg(c['items'], c.get('lb', c['id']))

    # xlsx から 参照 と編集可能テキスト欄を抜き出す
    xlsx_refs = {}     # {id: text (raw)}
    xlsx_seen = set()  # 参照欄を持つエンティティ（空欄含む）
    xlsx_fields = {}   # {id: {json_field: new_value}}
    wb = load_workbook(XLSX_PATH, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        id_col = next((i+1 for i,h in enumerate(hdrs) if h == 'id'), None)
        refs_col = next((i+1 for i,h in enumerate(hdrs) if h and '参照' in str(h)), None)
        stype = _detect_sheet_type(hdrs)
        # フィールド列マップ {col_idx: json_field_name}
        field_cols = {}
        if stype and stype in FIELD_MAPS:
            for i, h in enumerate(hdrs, 1):
                if h in FIELD_MAPS[stype]:
                    field_cols[i] = FIELD_MAPS[stype][h]
        if not id_col: continue
        for r in range(2, ws.max_row+1):
            eid = ws.cell(row=r, column=id_col).value
            if not eid: continue
            eid = str(eid).strip()
            # 参照
            if refs_col:
                xlsx_seen.add(eid)
                v = ws.cell(row=r, column=refs_col).value
                if v and str(v).strip():
                    xlsx_refs[eid] = str(v).strip()
            # 他のフィールド
            for col, jf in field_cols.items():
                v = ws.cell(row=r, column=col).value
                if v is None: continue
                v = str(v).strip()
                if v == '': continue
                xlsx_fields.setdefault(eid, {})[jf] = v

    # パース・検証
    parsed = {}      # {id: [refs]}
    errors = []
    for eid, text in xlsx_refs.items():
        if eid not in gid:
            errors.append(f'[元IDがJSONに存在しない] {eid}: `{text}`')
            continue
        try:
            refs = parse_refs(text, gid, name_to_ids)
        except ValueError as e:
            errors.append(f'[書式不正] {eid}: {e}')
            continue
        for r in refs:
            if r['id'] == eid:
                errors.append(f'[自己参照] {eid} → {r["id"]}')
            elif r['id'] not in gid:
                errors.append(f'[参照先ID不在] {eid} → {r["id"]}')
        parsed[eid] = refs

    # 差分計算（既存 JSON の refs と比較）
    add, change, remove = [], [], []
    all_entities = []
    def walk(arr, label):
        for p in arr:
            all_entities.append((p, label))
    for k, lb in people_cats: walk(d[k], lb)
    walk(d['noh'], '能曲')
    for c in d['cats']: walk(c['items'], c.get('lb', c['id']))

    for p, lb in all_entities:
        eid = p.get('id')
        if eid not in xlsx_seen:
            continue  # xlsx に欄が無いエンティティはそのまま温存
        cur = p.get('refs')
        new = parsed.get(eid)  # 空欄なら None
        if cur is None and new is None:
            pass
        elif cur is None and new is not None:
            add.append((eid, p.get('n'), lb, new))
        elif cur is not None and new is None:
            remove.append((eid, p.get('n'), lb, cur))
        elif cur != new:
            change.append((eid, p.get('n'), lb, cur, new))

    # フィールド変更の差分計算
    field_diffs = []  # (eid, name, label, field, old, new)
    by_id = {p.get('id'): (p, lb) for p, lb in all_entities}
    for eid, fields in xlsx_fields.items():
        if eid not in by_id: continue
        p, lb = by_id[eid]
        for field, new_val in fields.items():
            old_val = p.get(field)
            old_str = '' if old_val is None else str(old_val)
            if old_str != new_val:
                field_diffs.append((eid, p.get('n'), lb, field, old_val, new_val))

    # ----- レポート -----
    print('=' * 60)
    print(f'xlsx: {XLSX_PATH}')
    print(f'json: {JSON_PATH}  (version={d.get("version")})')
    print('=' * 60)
    print(f'xlsx の参照入力: {len(xlsx_refs)} エンティティ')
    print(f'検証エラー: {len(errors)}')
    for e in errors[:20]:
        print(f'  ✗ {e}')
    if len(errors) > 20:
        print(f'  ... ほか {len(errors)-20} 件')
    print()
    print(f'差分: 追加 {len(add)} / 変更 {len(change)} / 削除 {len(remove)}')

    def fmt(refs):
        return '; '.join(r['id'] + (f'({r["role"]})' if r.get('role') else '') for r in refs)

    if add:
        print('\n[追加]')
        for eid, n, lb, refs in add:
            print(f'  + [{lb}] {eid} ({n}) refs = {fmt(refs)}')
    if change:
        print('\n[変更]')
        for eid, n, lb, cur, new in change:
            print(f'  ~ [{lb}] {eid} ({n})')
            print(f'      旧: {fmt(cur)}')
            print(f'      新: {fmt(new)}')
    if remove:
        print('\n[削除]')
        for eid, n, lb, cur in remove:
            print(f'  - [{lb}] {eid} ({n}) refs（{fmt(cur)}）を削除')

    # フィールド変更レポート
    if field_diffs:
        print(f'\nフィールド変更: {len(field_diffs)} 件')
        # フィールドごとにグルーピング
        by_field = {}
        for d_ in field_diffs:
            by_field.setdefault(d_[3], []).append(d_)
        for field in sorted(by_field):
            print(f'\n  [{field}]')
            for eid, n, lb, _, old, new in by_field[field]:
                # 長文は冒頭のみ表示
                def trim(s):
                    s = '' if s is None else str(s)
                    return s if len(s) <= 60 else s[:60] + '…'
                print(f'    ~ [{lb}] {eid} ({n})')
                print(f'        旧: {trim(old)}')
                print(f'        新: {trim(new)}')

    if errors:
        print('\n⚠️  エラーがあるため書き込みは行いません。xlsx を修正してください。', file=sys.stderr)
        sys.exit(1)

    if not (add or change or remove or field_diffs):
        print('\n差分なし。何もしません。')
        return

    if not args.write:
        print('\n（ドライラン）--write を付けると実際に反映します。')
        return

    # ----- 反映 -----
    # バックアップ
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    bak = f'{JSON_PATH}.bak_{ts}'
    shutil.copy(JSON_PATH, bak)
    print(f'\nbackup: {bak}')

    # JSON 更新: refs
    for p, lb in all_entities:
        eid = p.get('id')
        if eid not in xlsx_seen: continue
        new = parsed.get(eid)
        if new:
            p['refs'] = new
        elif 'refs' in p:
            del p['refs']

    # JSON 更新: その他のテキストフィールド
    for eid, n, lb, field, old, new in field_diffs:
        if eid in by_id:
            by_id[eid][0][field] = new

    if not args.no_version_bump:
        old_ver = d.get('version', 0)
        d['version'] = old_ver + 1
        print(f'version: {old_ver} → {d["version"]}')

    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, separators=(',', ':'))

    print(f'\n✓ 書き込み完了: {JSON_PATH}')
    print('\n次の手順:')
    print('  1. index.html の DATA_VERSION_FALLBACK を新しい version に同期')
    print('  2. mobile/sw.js の CACHE を jhm-vNN に +1')
    print('  3. git add / commit / push')

if __name__ == '__main__':
    main()
