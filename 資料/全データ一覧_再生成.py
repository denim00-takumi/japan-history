#!/usr/bin/env python3
"""Generate 資料/全データ一覧_点検用.xlsx from japan_history_data.json.

- 各エンティティ・シートに「🔗 参照」列を追加（手動入力）
- 既存の xlsx があれば 参照 セルを保持してマージ
- 入力後に再実行すると「🔗 参照一覧」「⚠ 参照エラー」シートが自動更新される
"""
import json, re, os, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = '/Users/TN/Downloads/日本歴史年表'
SRC = os.path.join(ROOT, 'japan_history_data.json')
OUT = os.path.join(ROOT, '資料/全データ一覧_点検用.xlsx')

with open(SRC, 'r', encoding='utf-8') as f:
    d = json.load(f)

ERA_NAME = {
 'jomon':'縄文','yayoi':'弥生','kofun':'古墳','asuka':'飛鳥','nara':'奈良',
 'heian':'平安','kamakura':'鎌倉','nanboku':'南北朝','muromachi':'室町',
 'sengoku':'戦国','azuchi':'安土桃山','edo':'江戸','meiji':'明治','taisho':'大正',
 'showa':'昭和','heisei':'平成','reiwa':'令和'
}
def eraname(e): return ERA_NAME.get(e, e or '')

def yrtxt(v):
    if v is None or v == '': return ''
    try:
        n = int(v)
    except: return str(v)
    if n == 9999: return '存続中'
    if n < 0: return f'BC{-n}'
    return str(n)

HIRA = re.compile(r'^[぀-ゟー・\s]+$')

# ----- フォント既定値 -----
DEFAULT_FONTS = {
    'base':    {'name': 'Yu Gothic', 'size': 11},
    'header':  {'name': 'Yu Gothic', 'size': 11},
    'title':   {'name': 'Yu Gothic', 'size': 14},
    'mono':    {'name': 'Menlo',     'size': 10},
}
DEFAULT_ROW_HEIGHT = 22

# ----- ユーザフォント自動検出 -----
# 既存 xlsx のフォント/サイズ/行高を読み取り、再生成時に同じ体裁を保つ。
# Excel でフォントを変更して保存 → 次回再生成でその変更が引き継がれる。
def _detect_user_styling(path):
    out = {'fonts': dict(DEFAULT_FONTS), 'row_height': DEFAULT_ROW_HEIGHT}
    if not os.path.exists(path):
        return out
    try:
        from openpyxl import load_workbook as _lw
        wb_old = _lw(path)
        def take(sn, r, c):
            if sn not in wb_old.sheetnames: return None
            ws = wb_old[sn]
            if r > ws.max_row or c > ws.max_column: return None
            cell = ws.cell(row=r, column=c)
            if cell.font and cell.font.name:
                return {'name': cell.font.name, 'size': float(cell.font.size) if cell.font.size else None}
            return None
        # サマリー A1 → TITLE
        t = take('サマリー', 1, 1)
        if t: out['fonts']['title'] = t
        # 👤 人物 を参照（必ず存在）
        h = take('👤 人物', 1, 3)
        if h: out['fonts']['header'] = h
        b = take('👤 人物', 2, 3)
        if b: out['fonts']['base'] = b
        m = take('👤 人物', 2, 2)
        if m: out['fonts']['mono'] = m
        # データ行の行高
        if '👤 人物' in wb_old.sheetnames:
            ws = wb_old['👤 人物']
            rd = ws.row_dimensions.get(2)
            if rd and rd.height:
                out['row_height'] = float(rd.height)
    except Exception as e:
        print(f'warning: ユーザフォント検出失敗（既定値を使用）: {e}')
    return out

_user = _detect_user_styling(OUT)
_F = _user['fonts']
ROW_H = _user['row_height']

# ----- スタイル（ユーザ設定を反映） -----
FONT_BASE   = Font(name=_F['base']['name'],   size=_F['base']['size']   or 11)
FONT_HEADER = Font(name=_F['header']['name'], size=_F['header']['size'] or 11, bold=True, color='FFFFFF')
FONT_TITLE  = Font(name=_F['title']['name'],  size=_F['title']['size']  or 14, bold=True)
FONT_MONO   = Font(name=_F['mono']['name'],   size=_F['mono']['size']   or 10)
FONT_LINK   = Font(name=_F['mono']['name'],   size=_F['mono']['size']   or 10, color='1F6FB8')
FONT_NOTE   = Font(name=_F['base']['name'],   size=(_F['base']['size'] or 11) - 1, italic=True, color='6C757D')
FONT_ERR    = Font(name=_F['base']['name'],   size=_F['base']['size']   or 11, bold=True, color='B02A37')

print(f'フォント: base={_F["base"]["name"]} {_F["base"]["size"]}pt / mono={_F["mono"]["name"]} {_F["mono"]["size"]}pt / row高={ROW_H}')

FILL_HEADER = PatternFill('solid', start_color='2F3E55')
FILL_HEADER_REF = PatternFill('solid', start_color='1F6FB8')
FILL_ZEBRA = PatternFill('solid', start_color='F5F7FA')
FILL_REF = PatternFill('solid', start_color='EAF3FB')
FILL_WARN = PatternFill('solid', start_color='FFF3CD')
FILL_ERR = PatternFill('solid', start_color='F8D7DA')
FILL_OK = PatternFill('solid', start_color='D1E7DD')

THIN = Side(style='thin', color='C8CED5')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_HEAD = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

# ============== 既存 xlsx から 参照 セルを取得 ==============
existing_refs = {}  # {entity_id: refs_text}
if os.path.exists(OUT):
    try:
        wb_old = load_workbook(OUT, data_only=True)
        for sn in wb_old.sheetnames:
            ws_old = wb_old[sn]
            headers = [ws_old.cell(row=1, column=c).value for c in range(1, ws_old.max_column + 1)]
            id_col_idx = None
            refs_col_idx = None
            for i, h in enumerate(headers, 1):
                if h == 'id': id_col_idx = i
                if h and '参照' in str(h): refs_col_idx = i
            if not id_col_idx or not refs_col_idx:
                continue
            for r in range(2, ws_old.max_row + 1):
                eid = ws_old.cell(row=r, column=id_col_idx).value
                ref = ws_old.cell(row=r, column=refs_col_idx).value
                if eid and ref and str(ref).strip():
                    existing_refs[str(eid).strip()] = str(ref).strip()
        if existing_refs:
            print(f'preserved {len(existing_refs)} 参照 セル from existing xlsx')
    except Exception as e:
        print(f'warning: could not read existing xlsx: {e}', file=sys.stderr)

# ============== グローバル ID／名前 索引 ==============
# gid: {id: (name, category_label)}
# name_to_ids: {name: [(id, category_label), ...]}  ← 名前→ID 解決用
gid = {}
name_to_ids = {}
def register(arr, label):
    for p in arr:
        i = p.get('id')
        if i:
            gid[i] = (p.get('n', ''), label)
            nm = p.get('n', '')
            if nm:
                name_to_ids.setdefault(nm, []).append((i, label))

people_cats = [('persons','人物'),('emperors','天皇'),('rulers','為政者'),
               ('religious','宗教者'),('pms','総理大臣'),('rakugo','落語家'),
               ('kodan','講談師')]
for k, lb in people_cats: register(d[k], lb)
register(d['noh'], '能曲')
for c in d['cats']:
    register(c['items'], c.get('lb', c['id']))

ID_PATTERN = re.compile(r'^[A-Za-z0-9_]+$')

def resolve_token(token):
    """
    トークン（ID または 名前）を canonical ID に解決する。
    Returns (id|None, error_msg|None)
    """
    s = token.strip()
    if not s:
        return None, '空トークン'
    # 1) ID として存在すれば採用
    if s in gid:
        return s, None
    # 2) 名前として一意なら採用
    if s in name_to_ids:
        ids = name_to_ids[s]
        if len(ids) == 1:
            return ids[0][0], None
        choices = ', '.join(f'{i}({lb})' for i, lb in ids)
        return None, f'名前重複: `{s}` → {choices} のいずれか'
    # 3) ASCII っぽいなら ID 不在
    if ID_PATTERN.match(s):
        return None, f'ID不在: `{s}`'
    # 4) 名前未一致（近似候補を出す）
    suggestions = [nm for nm in name_to_ids if s in nm or nm in s][:3]
    if suggestions:
        return None, f'名前未一致: `{s}`（近似: {", ".join(suggestions)}）'
    return None, f'名前未一致: `{s}`'

# ============== Workbook 構築 ==============
wb = Workbook()
wb.remove(wb.active)

def write_header(ws, headers, row=1, ref_header_idx=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_REF if (ref_header_idx and i == ref_header_idx) else FILL_HEADER
        c.alignment = ALIGN_HEAD
        c.border = BORDER
    ws.row_dimensions[row].height = 28

def write_row(ws, row_idx, values, mono_cols=(), num_cols=(), zebra=False,
              ref_col_idx=None, link_cols=()):
    fill = FILL_ZEBRA if zebra else None
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        if i in link_cols:
            c.font = FONT_LINK
        elif i in mono_cols:
            c.font = FONT_MONO
        else:
            c.font = FONT_BASE
        if i in num_cols:
            c.alignment = ALIGN_RIGHT
        elif i == 1:
            c.alignment = ALIGN_CENTER
        else:
            c.alignment = ALIGN_LEFT
        c.border = BORDER
        if i == ref_col_idx:
            c.fill = FILL_REF
            c.font = FONT_LINK if v else FONT_MONO
        elif fill:
            c.fill = fill

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws, first_data_row=2, last_col=1):
    ws.freeze_panes = ws.cell(row=first_data_row, column=1)
    last_col_letter = get_column_letter(last_col)
    ws.auto_filter.ref = f'A1:{last_col_letter}{ws.max_row}'

# ============== サマリー ==============
ws = wb.create_sheet('サマリー')
ws['A1'] = '日本歴史年表 — 全データ一覧（点検・管理用）'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:C1')
ws['A2'] = f'出典: japan_history_data.json  version {d.get("version")}'
ws['A2'].font = FONT_BASE
ws['A3'] = '生成日: 2026-06-29  ／ 参照欄の書式: id(役割); id(役割); ...'
ws['A3'].font = FONT_NOTE

write_header(ws, ['カテゴリ', 'ID', '件数'], row=5)
summary = [
    ('👤 人物', 'persons', len(d['persons'])),
    ('👑 天皇', 'emperors', len(d['emperors'])),
    ('🛡 為政者', 'rulers', len(d['rulers'])),
    ('🛐 宗教者', 'religious', len(d['religious'])),
    ('🏛 総理大臣', 'pms', len(d['pms'])),
    ('🎤 落語家', 'rakugo', len(d['rakugo'])),
    ('📜 講談師', 'kodan', len(d['kodan'])),
    ('🎭 能曲', 'noh', len(d['noh'])),
]
for c in d['cats']:
    summary.append((c.get('lb', ''), c['id'], len(c['items'])))

for i, (lb, idv, cnt) in enumerate(summary, 6):
    zebra = (i % 2 == 0)
    for col, val in enumerate([lb, idv, cnt], 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = FONT_BASE if col != 2 else FONT_MONO
        cell.border = BORDER
        cell.alignment = ALIGN_RIGHT if col == 3 else ALIGN_LEFT
        if zebra: cell.fill = FILL_ZEBRA

total_row = 6 + len(summary)
ws.cell(row=total_row, column=1, value='合計').font = Font(name='Yu Gothic', size=11, bold=True)
ws.cell(row=total_row, column=3, value=f'=SUM(C6:C{total_row-1})').font = Font(name='Yu Gothic', size=11, bold=True)
for col in range(1, 4):
    c = ws.cell(row=total_row, column=col)
    c.border = BORDER
    c.fill = FILL_OK
    if col == 3: c.alignment = ALIGN_RIGHT
    if col == 1: c.alignment = ALIGN_LEFT
autosize(ws, [22, 18, 10])

# ============== 人物系シート共通 ==============
def make_people_sheet(name, arr):
    ws = wb.create_sheet(name)
    headers = ['✓', 'id', '名前', 'ふりがな', '生', '没', '時代', '役/分類', '概要', 'ピン数', '🔗 参照']
    ref_col = len(headers)
    write_header(ws, headers, ref_header_idx=ref_col)
    for i, p in enumerate(arr, 2):
        eid = p['id']
        ref = existing_refs.get(eid, '')
        write_row(ws, i,
                  ['', eid, p.get('n', ''), p.get('fg', ''),
                   yrtxt(p.get('s', '')), yrtxt(p.get('e', '')),
                   eraname(p.get('era', '')),
                   p.get('role', '') or p.get('l', '') or '',
                   p.get('desc', '') or '',
                   len(p.get('locs', [])),
                   ref],
                  mono_cols=(2,), num_cols=(10,), zebra=(i % 2 == 0),
                  ref_col_idx=ref_col)
        ws.row_dimensions[i].height = ROW_H
    autosize(ws, [4, 18, 22, 22, 8, 8, 10, 22, 50, 6, 30])
    freeze_and_filter(ws, first_data_row=2, last_col=len(headers))

make_people_sheet('👤 人物', d['persons'])
make_people_sheet('👑 天皇', d['emperors'])
make_people_sheet('🛡 為政者', d['rulers'])
make_people_sheet('🛐 宗教者', d['religious'])
make_people_sheet('🏛 総理大臣', d['pms'])
make_people_sheet('🎤 落語家', d['rakugo'])
make_people_sheet('📜 講談師', d['kodan'])

# ============== 能曲 ==============
ws = wb.create_sheet('🎭 能曲')
headers = ['✓', 'id', '曲名', 'ふりがな', '番', '分類', 'シテ類型', 'シテ', 'ワキ',
           '出典', '時代', '作者', '宗教', '注記', '🔗 参照']
ref_col = len(headers)
write_header(ws, headers, ref_header_idx=ref_col)
for i, n in enumerate(d['noh'], 2):
    eid = n['id']
    ref = existing_refs.get(eid, '')
    write_row(ws, i,
              ['', eid, n.get('n', ''), n.get('fg', ''),
               n.get('act', ''), n.get('cls', ''), n.get('shite_type', ''),
               n.get('shite', ''), n.get('waki', ''), n.get('source', ''),
               n.get('era', ''), n.get('author', ''), n.get('religion', ''),
               n.get('notes', ''), ref],
              mono_cols=(2,), num_cols=(5,), zebra=(i % 2 == 0),
              ref_col_idx=ref_col)
    ws.row_dimensions[i].height = ROW_H
autosize(ws, [4, 18, 20, 20, 5, 20, 12, 20, 20, 18, 14, 16, 14, 30, 30])
freeze_and_filter(ws, first_data_row=2, last_col=len(headers))

# ============== 五番立 ==============
ws = wb.create_sheet('五番立')
headers = ['番', '通称', 'ラベル', '概要', '色']
write_header(ws, headers)
for i, k in enumerate(sorted(d['nohActs'].keys(), key=int), 2):
    a = d['nohActs'][k]
    write_row(ws, i, [int(k), a.get('name', ''), a.get('label', ''), a.get('desc', ''), a.get('color', '')],
              mono_cols=(5,), num_cols=(1,), zebra=(i % 2 == 0))
    ws.row_dimensions[i].height = 24
autosize(ws, [6, 14, 14, 60, 12])
freeze_and_filter(ws, first_data_row=2, last_col=len(headers))

# ============== cats（宗教・芸能・出来事） ==============
def make_cat_sheet(c):
    lb = c.get('lb', c['id'])
    name = lb[:30]
    ws = wb.create_sheet(name)
    headers = ['✓', 'id', '名称', 'ふりがな', '開始', '終了', '時代', '概要', '🔗 参照']
    ref_col = len(headers)
    write_header(ws, headers, ref_header_idx=ref_col)
    for i, it in enumerate(c['items'], 2):
        eid = it['id']
        ref = existing_refs.get(eid, '')
        write_row(ws, i,
                  ['', eid, it.get('n', ''), it.get('fg', ''),
                   yrtxt(it.get('s', '')), yrtxt(it.get('e', '')),
                   eraname(it.get('era', '')), it.get('desc', ''),
                   ref],
                  mono_cols=(2,), zebra=(i % 2 == 0),
                  ref_col_idx=ref_col)
        ws.row_dimensions[i].height = ROW_H
    autosize(ws, [4, 18, 22, 22, 8, 8, 10, 50, 30])
    freeze_and_filter(ws, first_data_row=2, last_col=len(headers))

for c in d['cats']:
    make_cat_sheet(c)

# ============== 参照 パース ==============
# 入力形式: token1(役割1); token2(役割2); token3
#   token は ID（saicho）または 名前（最澄）どちらでも可
TOKEN_RE = re.compile(r'^\s*([^()（）;；]+?)\s*(?:[（(]\s*([^)）]*?)\s*[)）])?\s*$')

flat_refs = []   # (from_id, from_name, from_cat, to_id, to_name, to_cat, role)
ref_errors = []  # (from_id, from_name, from_cat, 種別, 詳細)

for from_id, refs_text in existing_refs.items():
    if from_id not in gid:
        ref_errors.append((from_id, '?', '?', '元IDが存在しない', f'参照欄=`{refs_text}`'))
        continue
    from_name, from_cat = gid[from_id]
    tokens = re.split(r'[;；]', refs_text)
    for tok in tokens:
        s = tok.strip()
        if not s: continue
        m = TOKEN_RE.match(s)
        if not m:
            ref_errors.append((from_id, from_name, from_cat, '書式不正', f'トークン=`{s}`'))
            continue
        raw_target, role = m.group(1).strip(), (m.group(2) or '').strip()
        to_id, err = resolve_token(raw_target)
        if err:
            ref_errors.append((from_id, from_name, from_cat, '参照先解決失敗', err))
            continue
        if to_id == from_id:
            ref_errors.append((from_id, from_name, from_cat, '自己参照', f'to=`{to_id}`'))
            continue
        to_name, to_cat = gid[to_id]
        flat_refs.append((from_id, from_name, from_cat, to_id, to_name, to_cat, role))

# ============== 🔗 参照一覧 ==============
ws = wb.create_sheet('🔗 参照一覧')
ws['A1'] = '参照一覧（全シートの「🔗 参照」列をフラット展開）'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')
ws['A2'] = f'件数: {len(flat_refs)}'
ws['A2'].font = Font(name='Yu Gothic', size=11, bold=True)
ws['A3'] = '※ 各エンティティ シートの 🔗 参照 列に "id(役割); id; ..." 形式で入力すると、ここに反映されます。'
ws['A3'].font = FONT_NOTE
ws.merge_cells('A3:H3')

write_header(ws, ['元 id', '元 名前', '元 カテゴリ', '→', '先 id', '先 名前', '先 カテゴリ', '役割'], row=5)
if flat_refs:
    for i, (fi, fn, fc, ti, tn, tc, role) in enumerate(flat_refs, 6):
        zebra = (i % 2 == 0)
        cells = [fi, fn, fc, '→', ti, tn, tc, role]
        for col, v in enumerate(cells, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_MONO if col in (1, 5) else FONT_BASE
            c.alignment = ALIGN_CENTER if col == 4 else ALIGN_LEFT
            c.border = BORDER
            if zebra: c.fill = FILL_ZEBRA
        ws.row_dimensions[i].height = ROW_H
else:
    c = ws.cell(row=6, column=1, value='（まだ 参照 が入力されていません）')
    c.font = FONT_NOTE
    ws.merge_cells('A6:H6')

autosize(ws, [18, 22, 14, 4, 18, 22, 14, 18])
ws.freeze_panes = 'A6'

# ============== ⚠ 参照エラー ==============
ws = wb.create_sheet('⚠ 参照エラー')
ws['A1'] = '参照エラー'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:E1')
ws['A2'] = f'検出件数: {len(ref_errors)}'
ws['A2'].font = FONT_ERR if ref_errors else Font(name='Yu Gothic', size=11, bold=True, color='0F5132')
write_header(ws, ['元 id', '元 名前', '元 カテゴリ', '種別', '詳細'], row=4)
if ref_errors:
    for i, (fi, fn, fc, kind, detail) in enumerate(ref_errors, 5):
        for col, v in enumerate([fi, fn, fc, kind, detail], 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_MONO if col in (1, 5) else FONT_BASE
            c.fill = FILL_ERR
            c.alignment = ALIGN_LEFT
            c.border = BORDER
        ws.row_dimensions[i].height = ROW_H
else:
    c = ws.cell(row=5, column=1, value='エラーなし')
    c.font = Font(name='Yu Gothic', size=11, bold=True, color='0F5132')
    c.fill = FILL_OK
    ws.merge_cells('A5:E5')
autosize(ws, [18, 22, 14, 18, 36])
ws.freeze_panes = 'A5'

# ============== 📝 desc要編集 ==============
# refs 先のエンティティ名が「元エンティティの desc 文字列」に出現しないものを検出。
# 次回アップデートで文中インラインリンク化する際に、ユーザが desc を編集して
# 該当名を文中に含めるための作業リスト。
entity_by_id = {}
def reg_entity(arr):
    for p in arr:
        if p.get('id'): entity_by_id[p['id']] = p
for k, _ in people_cats: reg_entity(d[k])
reg_entity(d['noh'])
for c in d['cats']: reg_entity(c['items'])

desc_missing = []  # (from_id, from_name, from_cat, desc_excerpt, to_id, to_name, to_cat, hint)
for fi, fn, fc, ti, tn, tc, role in flat_refs:
    from_ent = entity_by_id.get(fi)
    if not from_ent: continue
    desc = from_ent.get('desc') or ''
    if tn and tn not in desc:
        excerpt = desc[:70] + ('…' if len(desc) > 70 else '')
        hint = f'desc 文中に「{tn}」を含めて編集してください'
        desc_missing.append((fi, fn, fc, excerpt, ti, tn, tc, hint))

ws = wb.create_sheet('📝 desc要編集')
ws['A1'] = 'desc要編集（refs先の名前が元 desc に出てこないもの）'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')
ws['A2'] = f'検出件数: {len(desc_missing)}'
ws['A2'].font = Font(name='Yu Gothic', size=11, bold=True,
                     color='B45F06' if desc_missing else '0F5132')
ws['A3'] = '※ 次回アップデートで「文中インラインリンク」に切り替える前に、ここが 0件 になるよう desc を編集してください。'
ws['A3'].font = FONT_NOTE
ws.merge_cells('A3:H3')

write_header(ws, ['元 id', '元 名前', '元 カテゴリ', '元 desc 冒頭', '先 id', '先 名前', '先 カテゴリ', '推奨対応'], row=5)
if desc_missing:
    for i, (fi, fn, fc, ex, ti, tn, tc, hint) in enumerate(desc_missing, 6):
        for col, v in enumerate([fi, fn, fc, ex, ti, tn, tc, hint], 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_MONO if col in (1, 5) else FONT_BASE
            c.fill = FILL_WARN
            c.alignment = ALIGN_LEFT
            c.border = BORDER
        ws.row_dimensions[i].height = 30
else:
    c = ws.cell(row=6, column=1, value='全て desc 内に名前が含まれています — インラインリンク移行 OK')
    c.font = Font(name='Yu Gothic', size=11, bold=True, color='0F5132')
    c.fill = FILL_OK
    ws.merge_cells('A6:H6')
autosize(ws, [18, 22, 14, 50, 18, 22, 14, 36])
ws.freeze_panes = 'A6'

# ============== 整合性チェック（既存） ==============
ws = wb.create_sheet('⚠ 整合性チェック')
ws['A1'] = '整合性チェック結果'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:E1')

issues = []
def check_fg(arr, label):
    for p in arr:
        if not p.get('fg'):
            issues.append(('ふりがな未登録', label, p.get('id'), p.get('n'), ''))
def check_fg_clean(arr, label):
    for p in arr:
        fg = p.get('fg', '')
        if fg and not HIRA.match(fg):
            issues.append(('ふりがな表記揺れ', label, p.get('id'), p.get('n'), f'fg="{fg}"'))
def check_desc(arr, label):
    for p in arr:
        if 'desc' in p and not p.get('desc'):
            issues.append(('概要(desc)空', label, p.get('id'), p.get('n'), ''))
def check_locs(arr, label):
    for p in arr:
        if 'locs' in p and not p.get('locs'):
            issues.append(('地図ピン(locs)空', label, p.get('id'), p.get('n'), ''))
def check_range(arr, label):
    for p in arr:
        try:
            si, ei = int(p.get('s')), int(p.get('e'))
            if ei != 9999 and si > ei:
                issues.append(('生没年逆転', label, p.get('id'), p.get('n'), f's={si} > e={ei}'))
        except: pass

for k, lb in people_cats:
    check_fg(d[k], lb); check_fg_clean(d[k], lb)
    check_desc(d[k], lb); check_locs(d[k], lb); check_range(d[k], lb)
check_fg(d['noh'], '能曲'); check_fg_clean(d['noh'], '能曲')
for c in d['cats']:
    check_fg(c['items'], c.get('lb', c['id']))
    check_fg_clean(c['items'], c.get('lb', c['id']))
    check_desc(c['items'], c.get('lb', c['id']))
    check_range(c['items'], c.get('lb', c['id']))

write_header(ws, ['種別', 'カテゴリ', 'id', '名前', '詳細'], row=3)
if issues:
    for i, it in enumerate(issues, 4):
        for col, v in enumerate(it, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_MONO if col == 3 else FONT_BASE
            c.border = BORDER
            c.alignment = ALIGN_LEFT
            c.fill = FILL_WARN
        ws.row_dimensions[i].height = ROW_H
else:
    c = ws.cell(row=4, column=1, value='問題は検出されませんでした')
    c.font = FONT_BASE
    c.fill = FILL_OK
    ws.merge_cells('A4:E4')
ws['A2'] = f'検出件数: {len(issues)}'
ws['A2'].font = Font(name='Yu Gothic', size=11, bold=True,
                     color='B45F06' if issues else '0F5132')
autosize(ws, [18, 14, 22, 22, 40])
ws.freeze_panes = 'A4'

wb.save(OUT)
print(f'wrote: {OUT}')
print(f'sheets: {wb.sheetnames}')
print(f'参照 入力済み: {len(existing_refs)} 件 / 展開: {len(flat_refs)} 件 / エラー: {len(ref_errors)} 件')
print(f'整合性 issues: {len(issues)}')
